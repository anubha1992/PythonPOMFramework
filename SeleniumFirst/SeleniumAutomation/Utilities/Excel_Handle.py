from openpyxl import load_workbook, Workbook,worksheet

def getExcelData(sheetName):
    workbook = Workbook()
    workbook = load_workbook('..\TestData\Data.xlsx')
    sheet = workbook[sheetName]
    rows = sheet.max_row
    cols = sheet.max_column
    cell_list = []
    whole_list = []
    for row in range(2,rows+1):
        for col in range(1,cols+1):
            val = sheet.cell(row, col).value
            print(val)
            cell_list.append(val)
        #whole_list.extend(cell_list)
        whole_list.append(list(cell_list))
        cell_list.clear()
    print(whole_list)
    return whole_list












